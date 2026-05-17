"""DeepSeek 平台在线数据获取 — Playwright 自动化脚本

用法: python fetch_online.py <status_file_path>
"""

import csv
import glob
import json
import os
import sys
import tempfile
import time
import zipfile
from datetime import datetime

STATUS_FILE = sys.argv[1] if len(sys.argv) > 1 else None
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USER_DATA_DIR = os.path.join(os.path.expanduser("~"), ".deepseek-browser-data")


def write_status(status, message="", **extra):
    if not STATUS_FILE:
        return
    data = {"status": status, "message": message, "updated_at": datetime.now().isoformat()}
    data.update(extra)
    tmp = STATUS_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, STATUS_FILE)


# ── 列名模糊匹配 ─────────────────────────────────────────────

# 目标列 → 可能的源列名（中英文），按优先级排列
AMOUNT_COLUMN_MAP = {
    "user_id":      ["user_id", "用户ID", "user id", "userid", "使用者", "用户"],
    "utc_date":     ["utc_date", "日期", "date", "utc date", "时间", "日期时间", "timestamp"],
    "model":        ["model", "模型", "model_name", "model name", "模型名称"],
    "api_key_name": ["api_key_name", "API Key名称", "key name", "key_name", "apikey名称", "api key别名", "api_key"],
    "api_key":      ["api_key", "API Key", "key", "apikey", "密钥", "api_key"],
    "type":         ["type", "类型", "token_type", "token type", "token类型"],
    "price":        ["price", "单价", "unit price", "unit_price", "价格"],
    "amount":       ["amount", "用量", "数量", "usage", "tokens", "token数", "消耗",
                     "prompt_tokens", "completion_tokens", "billed_tokens",
                     "total_tokens", "output_tokens", "input_tokens"],
}

COST_COLUMN_MAP = {
    "user_id":      ["user_id", "用户ID", "user id", "userid", "使用者", "用户"],
    "utc_date":     ["utc_date", "日期", "date", "utc date", "时间", "timestamp"],
    "model":        ["model", "模型", "model_name", "model name", "模型名称"],
    "wallet_type":  ["wallet_type", "钱包类型", "wallet type", "wallet", "类型", "type"],
    "cost":         ["cost", "花费", "费用", "金额", "总价", "amount", "用量", "billed_tokens"],
    "currency":     ["currency", "货币", "币种", "单位"],
}


def find_column_mapping(actual_columns, target_map):
    """根据实际列名找到最佳映射。返回 {target_col: actual_col} 或 None。"""
    mapping = {}
    actual_lower = [c.strip().lower() for c in actual_columns]
    actual_orig = [c.strip() for c in actual_columns]

    for target_col, candidates in target_map.items():
        found = None
        for candidate in candidates:
            candidate_lower = candidate.lower()
            for i, ac in enumerate(actual_lower):
                if candidate_lower == ac:
                    found = i
                    break
            if found is not None:
                break
        if found is not None:
            mapping[target_col] = actual_orig[found]

    return mapping


def read_excel_to_rows(filepath, column_map, all_columns):
    """从 Excel 文件中读取数据行。返回 list[dict]"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = [str(c).strip() if c else "" for c in next(rows_iter)]

    mapping = find_column_mapping(header, column_map)

    # 检查必需列
    missing = [k for k in all_columns if k not in mapping]
    if missing:
        wb.close()
        expected = ", ".join(all_columns)
        actual = ", ".join(header)
        raise ValueError(f"列名匹配失败。\n期望: {expected}\n实际: {actual}\n未匹配: {', '.join(missing)}")

    rows = []
    for row_data in rows_iter:
        row = {}
        for target_col in all_columns:
            actual_col = mapping.get(target_col)
            idx = header.index(actual_col) if actual_col in header else -1
            val = str(row_data[idx]) if idx >= 0 and row_data[idx] is not None else ""
            row[target_col] = val
        rows.append(row)

    wb.close()
    return rows


def read_csv_file(filepath):
    """从 CSV 文件中读取数据行并识别类型。返回 (rows, file_type)"""
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    if not rows:
        return [], None

    header = [str(c).strip() if c else "" for c in fieldnames]

    # 检查是否匹配 amount 格式
    am = find_column_mapping(header, AMOUNT_COLUMN_MAP)
    am_matched = sum(1 for k in ["model", "type"] if k in am)
    if am_matched >= 2:
        # 为每行添加匹配后的列名键
        mapped_rows = []
        for row in rows:
            mr = {}
            for target_col, source_col in am.items():
                mr[target_col] = row.get(source_col, "")
            mapped_rows.append(mr)
        return mapped_rows, "amount"

    # 检查是否匹配 cost 格式
    cm = find_column_mapping(header, COST_COLUMN_MAP)
    cm_matched = sum(1 for k in ["model", "cost"] if k in cm)
    if cm_matched >= 2:
        mapped_rows = []
        for row in rows:
            mr = {}
            for target_col, source_col in cm.items():
                mr[target_col] = row.get(source_col, "")
            mapped_rows.append(mr)
        return mapped_rows, "cost"

    print(f"[fetch_online] CSV 列名无法识别: {fieldnames}")
    return [], None


def identify_file_type(filepath):
    """根据列名判断文件类型: 'amount' | 'cost' | None"""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
    wb.close()

    # 检查是否匹配 amount 格式
    am = find_column_mapping(header, AMOUNT_COLUMN_MAP)
    am_matched = sum(1 for k in ["user_id", "model", "type"] if k in am)
    if am_matched >= 3:
        return "amount", header

    # 检查是否匹配 cost 格式
    cm = find_column_mapping(header, COST_COLUMN_MAP)
    cm_matched = sum(1 for k in ["user_id", "model", "cost"] if k in cm)
    if cm_matched >= 3:
        return "cost", header

    return None, header


def merge_csv(filename, fieldnames, new_rows):
    """合并新行到已有 CSV，按所有字段去重。原子写入。"""
    filepath = os.path.join(BASE_DIR, filename)
    existing_rows = []
    existing_set = set()

    if os.path.exists(filepath):
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                existing_rows.append(row)
                key = tuple(row.get(k, "") for k in fieldnames)
                existing_set.add(key)

    added = 0
    for row in new_rows:
        key = tuple(row.get(k, "") for k in fieldnames)
        if key not in existing_set:
            existing_rows.append(row)
            existing_set.add(key)
            added += 1

    tmp_fd, tmp_path = tempfile.mkstemp(dir=BASE_DIR, suffix=".csv")
    try:
        with os.fdopen(tmp_fd, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in existing_rows:
                writer.writerow({k: row.get(k, "") for k in fieldnames})
        os.replace(tmp_path, filepath)
    except BaseException:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise

    return added


def detect_month_from_data(rows, file_type):
    """从数据行推断月份，返回 'YYYY-M' 格式。"""
    for row in rows:
        date_str = row.get("utc_date", "")
        if date_str:
            try:
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
                    try:
                        dt = datetime.strptime(date_str[:10], fmt)
                        return f"{dt.year}-{dt.month}"
                    except ValueError:
                        continue
            except Exception:
                continue
    # fallback: 当前月份
    now = datetime.now()
    return f"{now.year}-{now.month}"


# ── 主流程 ───────────────────────────────────────────────────

def main():
    try:
        write_status("running", "正在启动浏览器...")

        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            os.makedirs(USER_DATA_DIR, exist_ok=True)

            context = p.chromium.launch_persistent_context(
                user_data_dir=USER_DATA_DIR,
                headless=False,
                channel="chrome",
                args=["--disable-blink-features=AutomationControlled"],
                no_viewport=True,
            )

            page = context.pages[0] if context.pages else context.new_page()

            write_status("running", "正在访问 DeepSeek 平台...")
            page.goto("https://platform.deepseek.com/usage", wait_until="domcontentloaded", timeout=30000)

            # ── 登录检测与等待（循环检测，直到确认已登录） ──
            login_deadline = time.time() + 300  # 最多等 5 分钟
            login_notified = False

            while time.time() < login_deadline:
                page.wait_for_timeout(2000)
                cur_url = page.url.lower()

                # 检测是否是登录页面：URL 含 login/signin，或页面包含登录表单
                url_is_login = any(kw in cur_url for kw in ["login", "signin", "sign_in", "sign-in", "signup", "sign-up", "register", "auth"])
                has_login_form = False
                try:
                    # 检查是否有登录相关表单元素
                    pw_fields = page.locator("input[type='password']")
                    code_fields = page.locator("input[type='text'][name*='code' i], input[name*='code' i], input[placeholder*='验证码' i], input[placeholder*='code' i], input[placeholder*='验证' i]")
                    phone_fields = page.locator("input[type='tel'], input[placeholder*='手机' i], input[placeholder*='phone' i], input[placeholder*='邮箱' i], input[placeholder*='email' i]")
                    login_btns = page.locator("button:has-text('登录'), button:has-text('Log in'), button:has-text('Sign in'), button:has-text('登 录'), button:has-text('Continue'), button:has-text('继续')")
                    # 有密码框 或 (有验证码框+手机框) 且 有登录按钮
                    has_pw = pw_fields.count() > 0
                    has_code_or_phone = (code_fields.count() > 0 or phone_fields.count() > 0)
                    has_btn = login_btns.count() > 0
                    has_login_form = (has_pw and has_btn) or (has_code_or_phone and has_btn)
                except Exception:
                    pass

                is_login_page = url_is_login or has_login_form

                if is_login_page:
                    if not login_notified:
                        write_status("awaiting_login", "请在浏览器中登录 DeepSeek 平台")
                        print("[fetch_online] 检测到登录页面，等待用户登录...")
                        login_notified = True
                    write_status("awaiting_login", "请在浏览器中登录 DeepSeek 平台")
                    continue

                # 不在登录页，检查是否在 usage 页面
                if "usage" in cur_url or "platform.deepseek.com" in cur_url:
                    # 等待页面稳定
                    page.wait_for_timeout(2000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass

                    # 二次确认：等待后再次检查是否还在 usage 页面
                    page.wait_for_timeout(1000)
                    cur_url2 = page.url.lower()

                    # 再次检查有没有登录表单
                    try:
                        pw = page.locator("input[type='password']").count()
                        code = page.locator("input[name*='code' i], input[placeholder*='验证码' i]").count()
                        btn = page.locator("button:has-text('登录'), button:has-text('Sign in')").count()
                        if (pw > 0 or code > 0) and btn > 0:
                            continue  # 仍在登录页
                    except Exception:
                        pass

                    if "usage" not in cur_url2 and "platform.deepseek.com" not in cur_url2:
                        # 不在 DeepSeek 平台，尝试导航
                        page.goto("https://platform.deepseek.com/usage", wait_until="domcontentloaded", timeout=30000)
                        continue

                    print("[fetch_online] 确认已登录，在 usage 页面")
                    write_status("running", "登录成功，正在获取数据...")
                    break

            else:
                write_status("error", "登录超时（5分钟），请重试")
                context.close()
                return 1

            # 查找"导出"按钮
            write_status("running", "正在查找导出按钮...")

            # 尝试滚动到"每月用量"区域
            try:
                monthly_section = page.locator("text=每月用量").first
                if monthly_section.count() > 0:
                    monthly_section.scroll_into_view_if_needed()
                    page.wait_for_timeout(500)
            except Exception:
                pass

            # 保存页面 HTML 用于调试
            html_path = os.path.join(BASE_DIR, ".fetch-debug.html")
            try:
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(page.content())
            except Exception:
                pass

            # 尝试多种选择器找到导出按钮
            export_btn = None
            selectors = [
                # Playwright text-based (最可靠)
                "button:has-text('导出')",
                "button:has-text('Export')",
                "button:has-text('下载')",
                "button:has-text('Download')",
                # getByRole 风格
                "[role='button']:has-text('导出')",
                # 链接类型
                "a:has-text('导出')",
                "a:has-text('导出数据')",
                "a:has-text('下载')",
                # span/div 作为按钮
                "span:has-text('导出')",
                "div:has-text('导出')",
                # 图标按钮 (可能 title 属性包含导出)
                "[title*='导出']",
                "[title*='export' i]",
                "[title*='下载']",
                "[title*='download' i]",
                "[aria-label*='导出']",
                "[aria-label*='export' i]",
                # 包含"每月用量"区域内的按钮
                "text=每月用量 ~ button",
                # XPath
                "//button[contains(text(),'导出')]",
                "//button[contains(text(),'Export')]",
                "//button[contains(text(),'下载')]",
                "//*[contains(@class,'export') or contains(@class,'download')]",
                "//*[contains(text(),'导出')][self::button or self::a or self::span or self::div]",
                # 任何包含"导出"文字的可见元素
                "text=导出 >> visible=true",
            ]

            for sel in selectors:
                try:
                    loc = page.locator(sel)
                    count = loc.count()
                    if count > 0:
                        export_btn = loc.first
                        # 确保可见和可点击
                        try:
                            export_btn.wait_for(state="visible", timeout=3000)
                        except Exception:
                            pass
                        print(f"[fetch_online] 找到导出按钮 ({count}个): {sel}")
                        break
                except Exception as e:
                    continue

            if not export_btn:
                # 最后尝试：用 page.getByText 查找
                try:
                    candidates = page.get_by_text("导出", exact=False)
                    if hasattr(candidates, 'count') and candidates.count() > 0:
                        # 找最近的 button 或可点击元素
                        export_btn = candidates.first
                        print(f"[fetch_online] 通过 getByText 找到: 导出")
                except Exception:
                    pass

            if not export_btn:
                # 诊断：列出页面上所有按钮和可点击元素
                diag_lines = []
                try:
                    all_btns = page.locator("button, a, [role='button'], span[class*='btn'], div[class*='btn']").all()
                    diag_lines.append(f"页面上共找到 {len(all_btns)} 个可能的按钮元素:")
                    for i, btn in enumerate(all_btns[:30]):
                        try:
                            txt = btn.inner_text().strip()[:60]
                            tag = btn.evaluate("el => el.tagName")
                            diag_lines.append(f"  [{i}] <{tag}> {txt}")
                        except Exception:
                            pass
                except Exception as e:
                    diag_lines.append(f"诊断失败: {e}")

                try:
                    # 查找任何包含 "导出" "export" "下载" "download" 文字的元素
                    for kw in ["导出", "Export", "export", "下载", "Download", "CSV"]:
                        matches = page.locator(f"text={kw}").all()
                        texts = []
                        for m in matches[:10]:
                            try:
                                texts.append(m.inner_text().strip()[:80])
                            except Exception:
                                pass
                        if texts:
                            diag_lines.append(f"含 '{kw}' 的元素: {texts}")
                except Exception:
                    pass

                diag_text = "\n".join(diag_lines)
                print(f"[fetch_online] 诊断信息:\n{diag_text}")

                # 保存诊断到文件
                diag_path = os.path.join(BASE_DIR, ".fetch-diag.txt")
                with open(diag_path, "w", encoding="utf-8") as f:
                    f.write(diag_text)

                # 截图 + 保存 HTML
                screenshot_path = os.path.join(BASE_DIR, ".fetch-debug.png")
                page.screenshot(path=screenshot_path, full_page=True)
                write_status("error",
                    f"未找到导出按钮，页面结构可能已变化。\n"
                    f"已保存截图: .fetch-debug.png\n"
                    f"已保存HTML: .fetch-debug.html\n"
                    f"已保存诊断: .fetch-diag.txt")
                context.close()
                return 1

            # 点击导出，等待下载
            write_status("running", "正在下载数据...")
            try:
                with page.expect_download(timeout=60000) as download_info:
                    export_btn.click()
                download = download_info.value

                # 保存下载的文件
                zip_path = os.path.join(tempfile.gettempdir(), f"deepseek-export-{int(time.time())}.zip")
                download.save_as(zip_path)
                print(f"[fetch_online] 下载完成: {zip_path}")
            except Exception as e:
                write_status("error", f"下载失败: {e}")
                context.close()
                return 1

            context.close()

        # ── 解压和处理 ──────────────────────────────────────
        write_status("running", "正在处理下载文件...")

        extract_dir = tempfile.mkdtemp(prefix="deepseek-")
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(extract_dir)

            # 查找 Excel 和 CSV 文件
            data_files = []
            for root, dirs, files in os.walk(extract_dir):
                for f in files:
                    if f.endswith((".xlsx", ".xls", ".csv")) and not f.startswith("~$"):
                        data_files.append(os.path.join(root, f))

            if not data_files:
                # 列出压缩包中有哪些文件
                file_list = []
                for root, dirs, files in os.walk(extract_dir):
                    for f in files:
                        file_list.append(f)
                write_status("error",
                    f"下载的压缩包中没有找到 Excel/CSV 文件。\n"
                    f"压缩包包含: {', '.join(file_list[:20])}")
                return 1

            print(f"[fetch_online] 找到 {len(data_files)} 个数据文件")

            total_added = 0
            processed_files = []

            for filepath in data_files:
                if filepath.endswith((".xlsx", ".xls")):
                    ftype, header = identify_file_type(filepath)
                    if ftype == "amount":
                        fieldnames = ["user_id", "utc_date", "model", "api_key_name", "api_key", "type", "price", "amount"]
                        rows = read_excel_to_rows(filepath, AMOUNT_COLUMN_MAP, fieldnames)
                    elif ftype == "cost":
                        fieldnames = ["user_id", "utc_date", "model", "wallet_type", "cost", "currency"]
                        rows = read_excel_to_rows(filepath, COST_COLUMN_MAP, fieldnames)
                    else:
                        print(f"[fetch_online] 跳过未知 Excel: {os.path.basename(filepath)} (列名: {header})")
                        continue
                else:
                    # CSV 文件：尝试识别类型
                    rows, ftype = read_csv_file(filepath)
                    if ftype == "amount":
                        fieldnames = ["user_id", "utc_date", "model", "api_key_name", "api_key", "type", "price", "amount"]
                    elif ftype == "cost":
                        fieldnames = ["user_id", "utc_date", "model", "wallet_type", "cost", "currency"]
                    else:
                        print(f"[fetch_online] 跳过未知 CSV: {os.path.basename(filepath)}")
                        continue

                month = detect_month_from_data(rows, ftype)
                filename = f"{ftype}-{month}.csv"
                added = merge_csv(filename, fieldnames, rows)
                total_added += added
                processed_files.append(f"{filename} (+{added}行)")
                print(f"[fetch_online] {filename}: 合并 {len(rows)} 行, 新增 {added} 行")

            if not processed_files:
                write_status("error", "未识别到可处理的数据文件。请检查导出格式。")
                return 1

            msg = "数据已更新 · " + ", ".join(processed_files)
            write_status("done", msg, ok=True, total_added=total_added)
            print(f"[fetch_online] 完成: {msg}")
            return 0

        finally:
            # 清理临时文件
            try:
                for root, dirs, files in os.walk(extract_dir, topdown=False):
                    for f in files:
                        os.unlink(os.path.join(root, f))
                    for d in dirs:
                        os.rmdir(os.path.join(root, d))
                os.rmdir(extract_dir)
            except Exception:
                pass
            try:
                if os.path.exists(zip_path):
                    os.unlink(zip_path)
            except Exception:
                pass

    except ImportError as e:
        write_status("error", "请先安装 Playwright: pip install playwright && playwright install chromium")
        print(f"[fetch_online] ImportError: {e}")
        return 1
    except Exception as e:
        write_status("error", str(e))
        print(f"[fetch_online] 错误: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
